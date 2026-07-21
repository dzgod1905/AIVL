"""Excel Reader unit (category: parser). Reads an uploaded Excel file into JSON.

When the user uploads a file the chat base64-encodes the .xlsx into the run input
(input["file_b64"]); this unit decodes it and reads sheets into one JSON:

    { file, source, sheet_count, total_rows,
      sheets: [ { name, columns, rows: [ {col: val, ...} ], row_count }, ... ] }

Row 1 of each sheet (or of the configured range) = header -> column names; later
rows -> dicts keyed by them. If no file was uploaded it returns a small fixed
sample so the builder can still be exercised. Deterministic code tool -> always
done on the first attempt.

Config params (declared in SPEC["params"], rendered generically by the builder):
  - take_input_from (required): source of the file. "session" = the file uploaded
    in the chat. The orchestrator engine reads this to decide whether to feed the
    run's session file into this step's input.
  - sheet: read only this sheet. Empty (default) reads every sheet. Errors if the
    named sheet is missing.
  - cells: A1-style range to restrict reading (e.g. "A1:C10"). Empty (default)
    reads the whole sheet.
"""
from __future__ import annotations

import base64
import binascii
import datetime as _dt
import io
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from shared import config as cfg
from shared.celery_app import celery_app
from node.base import run_step

NAME = "excel_reader"

# Self-describing catalog entry (see docs/adding-a-tool.md). id MUST equal NAME.
# `params` = user-configurable settings. Each descriptor is rendered generically
# by the builder (no per-tool UI code): the tool reads the value from
# payload["config"][key], falling back to `default`.
SPEC = {
    "id": NAME,
    "name": "Excel Reader",
    "category": "parser",
    "description": "Read an Excel file into rows. Configure which sheet and cell range to read.",
    "outputSchema": {"type": "object", "additionalProperties": True},
    "params": [
        {
            "key": "take_input_from",
            "label": "Take input from",
            "type": "enum",
            "options": ["session"],
            "default": "session",
            "required": True,
            "description": "Where the Excel file comes from. 'session' = the file uploaded in the chat.",
        },
        {
            "key": "sheet",
            "label": "Sheet name",
            "type": "string",
            "default": "",
            "placeholder": "empty = all sheets",
            "description": "Read only this sheet. Empty reads every sheet. Errors if the named sheet is missing.",
        },
        {
            "key": "cells",
            "label": "Cell range",
            "type": "string",
            "default": "",
            "placeholder": "empty = all cells, e.g. A1:C10",
            "description": "Restrict reading to this A1-style range. Empty reads the whole sheet.",
        },
    ],
}

_SAMPLE_SHEETS = [
    {
        "name": "Sheet1",
        "columns": ["id", "name", "amount"],
        "rows": [
            {"id": 1, "name": "Alpha", "amount": 100},
            {"id": 2, "name": "Beta", "amount": 250},
            {"id": 3, "name": "Gamma", "amount": 75},
        ],
        "row_count": 3,
    }
]


class _ExcelError(Exception):
    """A config/user error to surface verbatim (bad sheet name or range)."""


def _cell(v: Any) -> Any:
    """Coerce a cell value to something JSON-serializable."""
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        return v.isoformat()
    return v


def _bounds(cells: str) -> tuple[int, int, int, int] | None:
    """Parse an A1 range into (min_row, max_row, min_col, max_col) for iter_rows.
    Empty -> None (read the whole sheet). Raises _ExcelError on a bad range."""
    cells = (cells or "").strip()
    if not cells:
        return None
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cells)
    except (ValueError, TypeError) as exc:
        raise _ExcelError(f"invalid cell range {cells!r}: {exc}") from exc
    if None in (min_col, min_row, max_col, max_row):
        # open-ended ranges like "A:C" leave row bounds None; not supported here
        raise _ExcelError(f"cell range {cells!r} must be bounded, e.g. A1:C10")
    return min_row, max_row, min_col, max_col


def _read_sheet(ws: Any, bounds: tuple[int, int, int, int] | None) -> dict[str, Any]:
    if bounds is None:
        rows_iter = ws.iter_rows(values_only=True)
    else:
        min_row, max_row, min_col, max_col = bounds
        rows_iter = ws.iter_rows(
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col,
            values_only=True,
        )
    header = next(rows_iter, None)
    if header is None:
        return {"name": ws.title, "columns": [], "rows": [], "row_count": 0}

    columns = [str(h) if h is not None else f"col{i + 1}" for i, h in enumerate(header)]
    rows: list[dict[str, Any]] = []
    for r in rows_iter:
        if r is None or all(c is None for c in r):
            continue  # skip blank rows
        rows.append({columns[i]: _cell(r[i]) for i in range(min(len(columns), len(r)))})
    return {"name": ws.title, "columns": columns, "rows": rows, "row_count": len(rows)}


def _parse(raw: bytes, sheet: str, cells: str) -> list[dict[str, Any]]:
    bounds = _bounds(cells)
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        sheet = (sheet or "").strip()
        if sheet:
            if sheet not in wb.sheetnames:
                raise _ExcelError(
                    f"sheet {sheet!r} not found; available: {', '.join(wb.sheetnames)}"
                )
            names = [sheet]
        else:
            names = wb.sheetnames
        return [_read_sheet(wb[name], bounds) for name in names]
    finally:
        wb.close()


def _pack(file: str, source: str, sheets: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "file": file,
        "source": source,
        "sheet_count": len(sheets),
        "total_rows": sum(s["row_count"] for s in sheets),
        "sheets": sheets,
    }


def _build_output(input_obj: dict[str, Any], conf: dict[str, Any], attempt: int) -> dict[str, Any]:
    name = input_obj.get("file") or "sample.xlsx"
    b64 = input_obj.get("file_b64")
    sheet = str(conf.get("sheet", "") or "")
    cells = str(conf.get("cells", "") or "")

    if not b64:
        return _pack("sample.xlsx", "sample", _SAMPLE_SHEETS)

    # Reject oversized uploads before decoding/parsing (zip-bomb / OOM guard).
    # base64 inflates ~4/3, so pre-check the encoded length too.
    if len(b64) > cfg.MAX_XLSX_BYTES * 4 // 3 + 4:
        out = _pack(name, "error", [])
        out["error"] = f"file too large: exceeds {cfg.MAX_XLSX_BYTES} bytes"
        return out

    try:
        raw = base64.b64decode(b64, validate=True)
        if len(raw) > cfg.MAX_XLSX_BYTES:
            out = _pack(name, "error", [])
            out["error"] = f"file too large: {len(raw)} bytes > {cfg.MAX_XLSX_BYTES}"
            return out
        return _pack(name, "upload", _parse(raw, sheet, cells))
    except _ExcelError as exc:  # bad sheet name / range -> surface to the user
        out = _pack(name, "error", [])
        out["error"] = str(exc)
        return out
    except (binascii.Error, ValueError) as exc:
        out = _pack(name, "error", [])
        out["error"] = f"invalid base64: {exc}"
        return out
    except Exception as exc:  # not an xlsx / corrupt
        out = _pack(name, "error", [])
        out["error"] = f"cannot read excel: {exc}"
        return out


@celery_app.task(name="node.excel_reader")
def run(payload: dict[str, Any]) -> dict[str, Any]:
    return run_step(NAME, payload, _build_output, always_done=True)
